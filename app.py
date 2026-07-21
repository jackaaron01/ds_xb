import os, io, json, glob, sys, socket, threading, webbrowser
from datetime import datetime

from flask import Flask, render_template, request, jsonify, send_file
from PIL import Image, ImageDraw, ImageFont, ImageOps
from openpyxl import load_workbook

app = Flask(__name__)

if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_FILE = os.path.join(APP_DIR, "config.json")
FONT_CACHE_FILE = os.path.join(APP_DIR, "fonts_cache.json")
TEMPLATE_DIR = os.path.join(APP_DIR, "template")
OUTPUT_DIR = os.path.join(APP_DIR, "output")
FONT_DIR = "C:/Windows/Fonts"

os.makedirs(TEMPLATE_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

CN_FONT_KEYWORDS = [
    "simkai", "simhei", "simfang", "simsun", "simyou", "simli",
    "stkaiti", "stxingka", "stliti", "stsong", "stzhongs",
    "stxihei", "stcaiyun", "sthupo", "stfangso", "stxinwei",
    "msyh", "msjh", "msyi", "hysw", "hanyi", "fz", "fangzheng",
    "mingliu", "deng", "harmonyos", "notosans", "notoserif",
    "sourcehan", "yugoth", "weiruanyahei",
]

DEFAULT_TEXT_BOXES = [
    {"id": 1, "label": "姓名(大)", "text": "", "x": 350, "y": 600, "size": 80, "color": [180, 0, 0]},
    {"id": 2, "label": "姓名(小1)", "text": "", "x": 350, "y": 720, "size": 40, "color": [180, 0, 0]},
    {"id": 3, "label": "姓名(小2)", "text": "", "x": 350, "y": 780, "size": 40, "color": [180, 0, 0]},
    {"id": 4, "label": "学校", "text": "", "x": 350, "y": 860, "size": 50, "color": [180, 0, 0]},
    {"id": 5, "label": "日期", "text": "", "x": 350, "y": 960, "size": 36, "color": [100, 100, 100]},
]


def load_cfg():
    defs = {
        "bg_path": os.path.join(APP_DIR, "template", "bg.jpg"),
        "output_dir": OUTPUT_DIR,
        "font_path": "C:/Windows/Fonts/STXINGKA.TTF",
        "stroke_width": 0,
        "text_boxes": [dict(b) for b in DEFAULT_TEXT_BOXES],
        "output_size": "original",
        "output_format": "png",
        "quality": 95,
        "next_id": 6,
    }
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                content = f.read()
            if not content.strip():
                return dict(defs)
            cfg = json.loads(content)
        except (json.JSONDecodeError, OSError):
            return dict(defs)

        # 迁移旧版配置
        if "text_boxes" not in cfg:
            cfg["text_boxes"] = _migrate_old_config(cfg)
            cfg["next_id"] = max(b["id"] for b in cfg["text_boxes"]) + 1
        # 补齐缺失字段
        for k, v in defs.items():
            if k not in cfg:
                cfg[k] = v
        # 路径转绝对
        for key in ["bg_path", "output_dir"]:
            val = cfg.get(key, "")
            if val and not os.path.isabs(val):
                cfg[key] = os.path.join(APP_DIR, val)
        if not os.path.exists(cfg["bg_path"]):
            fb = os.path.join(APP_DIR, "template", "bg.jpg")
            if os.path.exists(fb):
                cfg["bg_path"] = fb
        return cfg
    return dict(defs)


def _migrate_old_config(cfg):
    """将旧版固定字段转换为text_boxes"""
    boxes = []
    tc = cfg.get("text_color", [180, 0, 0])
    dc = cfg.get("date_color", [100, 100, 100])
    bid = 1
    for key, label in [("name", "姓名(大)"), ("name2", "姓名(小1)"), ("name3", "姓名(小2)"),
                        ("school", "学校"), ("date", "日期")]:
        if key in cfg and isinstance(cfg[key], dict):
            item = cfg[key]
            boxes.append({
                "id": bid, "label": label, "text": "",
                "x": item.get("x", 350), "y": item.get("y", 600 + bid * 50),
                "size": item.get("size", 60), "color": dc if key == "date" else tc
            })
            bid += 1
    return boxes if boxes else [dict(b) for b in DEFAULT_TEXT_BOXES]


def save_cfg(cfg):
    to_save = dict(cfg)
    for key in ["bg_path", "output_dir"]:
        val = to_save.get(key, "")
        if val and os.path.isabs(val):
            try:
                to_save[key] = os.path.relpath(val, APP_DIR)
            except ValueError:
                pass
    # 原子写入：先写临时文件再替换，避免读写竞态
    tmp = CONFIG_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(to_save, f, ensure_ascii=False, indent=2)
    os.replace(tmp, CONFIG_FILE)


def resolve_font(size):
    cfg = load_cfg()
    fp = cfg.get("font_path", "")
    if fp and os.path.exists(fp):
        try:
            return ImageFont.truetype(fp, size)
        except Exception:
            pass
    candidates = [os.path.join(FONT_DIR, f) for f in
                  ["STXINGKA.TTF", "STKAITI.TTF", "simkai.ttf", "simhei.ttf",
                   "STLITI.TTF", "STZHONGS.TTF", "msyh.ttc", "simsun.ttc"]]
    for p in candidates:
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, size)
            except Exception:
                pass
    return ImageFont.load_default()


def draw_text(draw, pos, text, font, fill, stroke):
    x, y = pos
    draw.text((x, y), text, font=font, fill=fill)
    if stroke <= 0:
        return
    offsets_all = [
        [(-1, 0), (1, 0), (0, -1), (0, 1)],
        [(-1, -1), (-1, 1), (1, -1), (1, 1)],
        [(-2, 0), (2, 0), (0, -2), (0, 2)],
        [(-2, -1), (-2, 1), (2, -1), (2, 1)],
        [(-1, -2), (1, -2), (-1, 2), (1, 2)],
        [(-2, -2), (-2, 2), (2, -2), (2, 2)],
    ]
    for layer in range(stroke):
        if layer >= len(offsets_all):
            break
        for ox, oy in offsets_all[layer]:
            draw.text((x + ox, y + oy), text, font=font, fill=fill)


def draw_card(img, boxes, stroke):
    draw = ImageDraw.Draw(img)
    for box in boxes:
        text = box.get("text", "").strip()
        if not text:
            continue
        size = box.get("size", 60)
        x = box.get("x", 300)
        y = box.get("y", 500)
        color = tuple(box.get("color", [180, 0, 0]))
        font = resolve_font(size)
        bbox = draw.textbbox((0, 0), text, font=font)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        draw_text(draw, (x - tw // 2, y - th // 2), text, font, color, stroke)


# ===================== API 路由 =====================

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/config", methods=["GET", "POST"])
def api_config():
    if request.method == "GET":
        cfg = load_cfg()
        cfg["bg_exists"] = os.path.exists(cfg.get("bg_path", ""))
        return jsonify(cfg)
    else:
        cfg = load_cfg()
        data = request.get_json()
        if "text_boxes" in data:
            cfg["text_boxes"] = data["text_boxes"]
        for k in ["stroke_width", "output_size", "output_format", "quality", "font_path", "next_id"]:
            if k in data:
                cfg[k] = data[k]
        save_cfg(cfg)
        return jsonify({"status": "ok"})


@app.route("/api/fonts")
def api_fonts():
    refresh = request.args.get("refresh", "0")
    if refresh == "1":
        fonts = _scan_fonts()
    else:
        fonts = _load_fonts()
    return jsonify(fonts)


def _scan_fonts():
    fonts = []
    if os.path.isdir(FONT_DIR):
        for ext in ["*.ttf", "*.ttc", "*.otf"]:
            for fp in glob.glob(os.path.join(FONT_DIR, ext)):
                name = os.path.basename(fp)
                low = name.lower()
                if not any(kw in low for kw in CN_FONT_KEYWORDS):
                    continue
                try:
                    font = ImageFont.truetype(fp, 24)
                    dummy = Image.new("RGB", (100, 30), (255, 255, 255))
                    ImageDraw.Draw(dummy).text((2, 2), "中文", font=font, fill=(0, 0, 0))
                    if any(p[0] != 255 or p[1] != 255 or p[2] != 255 for p in dummy.getdata()):
                        fonts.append({"name": name, "path": fp})
                except Exception:
                    pass
    fonts.sort(key=lambda f: f["name"].lower())
    with open(FONT_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(fonts, f, ensure_ascii=False)
    return fonts


def _load_fonts():
    if os.path.exists(FONT_CACHE_FILE):
        try:
            with open(FONT_CACHE_FILE, "r", encoding="utf-8") as f:
                cached = json.load(f)
            if cached:
                return cached
        except Exception:
            pass
    return _scan_fonts()


@app.route("/api/preview")
def api_preview():
    cfg = load_cfg()
    bg = cfg.get("bg_path", "")
    if not os.path.exists(bg):
        return jsonify({"error": "背景图不存在"})

    boxes = [dict(b) for b in cfg.get("text_boxes", [])]
    # 用查询参数覆盖文本框文字
    for box in boxes:
        qk = f"t{box['id']}"
        if qk in request.args:
            box["text"] = request.args.get(qk, "")

    active_id = request.args.get("active", "1", type=int)
    stroke = cfg.get("stroke_width", 0)

    img = Image.open(bg).convert("RGB")
    w, h = img.size
    max_w = 500
    scale = max_w / w
    nw, nh = int(w * scale), int(h * scale)
    img_rz = img.resize((nw, nh), Image.LANCZOS).convert("RGBA")

    draw_rz = ImageDraw.Draw(img_rz)
    for box in boxes:
        text = box.get("text", "").strip()
        if not text:
            continue
        ss = max(6, int(box.get("size", 60) * scale))
        color = tuple(box.get("color", [180, 0, 0]))
        try:
            font = resolve_font(ss)
        except Exception:
            font = ImageFont.load_default()
        bbox = draw_rz.textbbox((0, 0), text, font=font)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        sx, sy = int(box["x"] * scale), int(box["y"] * scale)
        draw_text(draw_rz, (sx - tw // 2, sy - th // 2), text, font, color, max(0, int(stroke * scale)))

    overlay = Image.new("RGBA", (nw, nh), (0, 0, 0, 0))
    odraw = ImageDraw.Draw(overlay)
    for box in boxes:
        sx, sy = int(box["x"] * scale), int(box["y"] * scale)
        r = 4
        is_active = box["id"] == active_id
        fill = (0, 255, 0, 220) if is_active else (255, 0, 0, 200)
        odraw.ellipse([sx - r, sy - r, sx + r, sy + r], fill=fill)

    result = Image.alpha_composite(img_rz, overlay)
    buf = io.BytesIO()
    result.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")


@app.route("/api/font_preview/<path:fontpath>")
def api_font_preview(fontpath):
    text = request.args.get("text", "金榜题名")
    size = int(request.args.get("size", 28))
    bg_hex = request.args.get("bg", "1e2430")
    try:
        r = int(bg_hex[0:2], 16); g = int(bg_hex[2:4], 16); b = int(bg_hex[4:6], 16)
    except Exception:
        r, g, b = 30, 35, 45
    if not os.path.exists(fontpath):
        return jsonify({"error": "字体不存在"}), 404
    try:
        font = ImageFont.truetype(fontpath, size)
    except Exception:
        font = ImageFont.load_default()
    dummy = Image.new("RGBA", (1, 1))
    d = ImageDraw.Draw(dummy)
    bbox = d.textbbox((0, 0), text, font=font)
    tw, th = bbox[2] - bbox[0] + 8, bbox[3] - bbox[1] + 8
    pad = 6
    img = Image.new("RGBA", (tw + pad * 2, th + pad * 2), (r, g, b, 255))
    draw = ImageDraw.Draw(img)
    draw.text((pad - bbox[0], pad - bbox[1]), text, font=font, fill=(255, 215, 0, 255))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return send_file(buf, mimetype="image/png")


@app.route("/api/generate", methods=["POST"])
def api_generate():
    try:
        cfg = load_cfg()
        bg = cfg.get("bg_path", "")
        if not os.path.exists(bg):
            return jsonify({"error": "背景图不存在"}), 400

        data = request.get_json()
        boxes = data.get("text_boxes", cfg.get("text_boxes", []))

        img = Image.open(bg).convert("RGB")
        draw_card(img, boxes, cfg.get("stroke_width", 0))

        out_size = cfg.get("output_size", "original")
        if out_size != "original":
            w, h = map(int, out_size.split("x"))
            img = ImageOps.fit(img, (w, h), method=Image.LANCZOS)

        fmt = cfg.get("output_format", "png")
        ext = "jpg" if fmt == "jpg" else "png"
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fn = f"gaokao_{ts}.{ext}"
        save_path = os.path.join(cfg.get("output_dir", OUTPUT_DIR), fn)

        save_kwargs = {"quality": cfg.get("quality", 95)} if ext == "jpg" else {}
        buf = io.BytesIO()
        img.save(buf, format=ext.upper(), **save_kwargs)
        buf.seek(0)
        with open(save_path, "wb") as f:
            f.write(buf.read())
        return jsonify({"success": True, "filename": fn, "path": save_path})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500


@app.route("/api/download/<filename>")
def api_download(filename):
    cfg = load_cfg()
    return send_file(os.path.join(cfg.get("output_dir", OUTPUT_DIR), filename),
                     as_attachment=True, download_name=filename)


@app.route("/api/batch", methods=["POST"])
def api_batch():
    if "file" not in request.files:
        return jsonify({"error": "请上传Excel文件"}), 400

    file = request.files["file"]
    wb = load_workbook(file)
    ws = wb.active
    # 读取表头确定列映射
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    cfg = load_cfg()
    template_boxes = cfg.get("text_boxes", [])
    stroke = cfg.get("stroke_width", 0)

    results = []
    batch_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    idx = 0

    for row in rows:
        if not row or not row[0]:
            continue
        # 用表头匹配填充文本框
        txt_boxes = []
        for box in template_boxes:
            nb = dict(box)
            nb["text"] = ""
            txt_boxes.append(nb)
        for ci, cell in enumerate(row):
            if ci >= len(headers) or not cell:
                continue
            val = str(cell).strip()
            hl = headers[ci].lower()
            # 匹配标签名称
            for box in txt_boxes:
                label_low = box["label"].lower()
                if hl in label_low or label_low in hl:
                    box["text"] = val
                    break
        try:
            img = Image.open(cfg["bg_path"]).convert("RGB")
            draw_card(img, txt_boxes, stroke)
            out_size = cfg.get("output_size", "original")
            if out_size != "original":
                w, h = map(int, out_size.split("x"))
                img = ImageOps.fit(img, (w, h), method=Image.LANCZOS)
            fmt = cfg.get("output_format", "png")
            ext = "jpg" if fmt == "jpg" else "png"
            idx += 1
            ts = f"{batch_ts}_{idx:03d}"
            fn = f"gaokao_{ts}.{ext}"
            sp = os.path.join(cfg.get("output_dir", OUTPUT_DIR), fn)
            buf = io.BytesIO()
            img.save(buf, format=ext.upper(), **({"quality": cfg.get("quality", 95)} if ext == "jpg" else {}))
            buf.seek(0)
            with open(sp, "wb") as f:
                f.write(buf.read())
            name = str(row[0]).strip() if row else "unknown"
            results.append({"name": name, "success": True, "filename": fn})
        except Exception as e:
            results.append({"name": str(row[0]) if row else "?", "success": False, "error": str(e)})

    ok = sum(1 for r in results if r["success"])
    return jsonify({"total": len(results), "success": ok, "results": results})


@app.route("/api/upload_bg", methods=["POST"])
def api_upload_bg():
    if "file" not in request.files:
        return jsonify({"error": "请上传图片"}), 400
    file = request.files["file"]
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".jpg", ".jpeg", ".png", ".bmp"):
        return jsonify({"error": "仅支持 JPG/PNG/BMP"}), 400
    save_path = os.path.join(TEMPLATE_DIR, f"bg{ext}")
    file.save(save_path)
    cfg = load_cfg()
    cfg["bg_path"] = os.path.join("template", f"bg{ext}")
    save_cfg(cfg)
    return jsonify({"success": True, "filename": f"bg{ext}"})


@app.route("/api/bg_info")
def api_bg_info():
    cfg = load_cfg()
    bg = cfg.get("bg_path", "")
    if not os.path.exists(bg):
        return jsonify({"error": "无背景图"})
    img = Image.open(bg)
    return jsonify({"width": img.size[0], "height": img.size[1], "filename": os.path.basename(bg)})


@app.route("/api/getip")
def api_getip():
    hostname = socket.gethostname()
    ip = socket.gethostbyname(hostname)
    return jsonify({"ip": ip})


@app.route("/api/files")
def api_files():
    files = []
    out_dir = load_cfg().get("output_dir", OUTPUT_DIR)
    if os.path.isdir(out_dir):
        for fn in os.listdir(out_dir):
            fp = os.path.join(out_dir, fn)
            if os.path.isfile(fp) and fn.lower().endswith((".png", ".jpg", ".jpeg")):
                files.append({"name": fn, "size": os.path.getsize(fp),
                              "time": datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%m-%d %H:%M")})
    files.sort(key=lambda f: f["time"], reverse=True)
    return jsonify(files[:20])


@app.route("/api/config/reset", methods=["POST"])
def api_config_reset():
    if os.path.exists(CONFIG_FILE):
        os.remove(CONFIG_FILE)
    return jsonify({"status": "ok"})


@app.route("/api/preview_image/<filename>")
def api_preview_image(filename):
    cfg = load_cfg()
    fp = os.path.join(cfg.get("output_dir", OUTPUT_DIR), filename)
    if not os.path.exists(fp):
        return jsonify({"error": "文件不存在"}), 404
    return send_file(fp, mimetype="image/png" if fp.lower().endswith(".png") else "image/jpeg")


def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


def find_free_port(start=5000, max_port=5020):
    for port in range(start, max_port + 1):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.bind(("", port))
            s.close()
            return port
        except OSError:
            continue
    return start


if __name__ == "__main__":
    host = "0.0.0.0"
    port = find_free_port(5000, 5020)
    url = f"http://127.0.0.1:{port}"

    if not getattr(sys, 'frozen', False):
        app.run(host=host, port=port, debug=True)
    else:
        import logging
        logging.getLogger('werkzeug').setLevel(logging.WARNING)
        print("=" * 50)
        print("   [高考喜报自动生成系统 v3.0]")
        print("=" * 50)
        print(f"\n   本机访问: {url}")
        try:
            lip = get_local_ip()
            if lip and lip != "127.0.0.1":
                print(f"   局域网访问: http://{lip}:{port}")
        except Exception:
            pass
        print(f"\n   浏览器将自动打开，请勿关闭本窗口。")
        print(f"   按 Ctrl+C 可以退出程序。\n")
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()
        app.run(host=host, port=port, debug=False)
