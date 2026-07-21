import os
import json
from datetime import datetime

from PIL import Image, ImageDraw, ImageFont, ImageOps
from openpyxl import load_workbook

from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QLineEdit, QPushButton,
    QVBoxLayout, QHBoxLayout, QGridLayout, QFileDialog,
    QMessageBox, QSpinBox, QComboBox, QGroupBox, QStatusBar,
    QMainWindow, QScrollArea, QFrame, QSizePolicy, QProgressDialog
)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QPixmap, QImage, QFont, QPainter, QColor, QPen

BG_PATH = "template/bg.jpg"
OUTPUT_DIR = "output"
FONT_PATH = ""
CONFIG_FILE = "config.json"

DEFAULT_CONFIG = {
    "bg_path": BG_PATH,
    "output_dir": OUTPUT_DIR,
    "font_path": FONT_PATH,
    "text_color": [180, 0, 0],
    "date_color": [100, 100, 100],
    "stroke_width": 0,
    "name": {"x": 527, "y": 580, "size": 72, "align": "center"},
    "name2": {"x": 527, "y": 740, "size": 48, "align": "center"},
    "name3": {"x": 527, "y": 650, "size": 40, "align": "center"},
    "school": {"x": 527, "y": 860, "size": 50, "align": "center"},
    "date": {"x": 527, "y": 980, "size": 40, "align": "center"},
    "output_size": "original",
    "output_format": "png",
    "quality": 95
}

PRESET_SIZES = {
    "原始尺寸": "original",
    "朋友圈 (1080x1440)": "1080x1440",
    "公众号封面 (900x383)": "900x383",
    "A4打印 (2480x3508)": "2480x3508"
}


def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        # 补齐旧版配置缺少的字段
        for k, v in DEFAULT_CONFIG.items():
            if k not in cfg:
                cfg[k] = v
        return cfg
    return dict(DEFAULT_CONFIG)


def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def get_font(size):
    cfg = load_config()
    fp = cfg.get("font_path", "")
    if fp and os.path.exists(fp):
        return ImageFont.truetype(fp, size)
    try:
        return ImageFont.truetype("msyh.ttc", size)
    except Exception:
        try:
            return ImageFont.truetype("simhei.ttf", size)
        except Exception:
            return ImageFont.load_default()


def create_card(name, name2, name3, school, date, show_msg=True):
    cfg = load_config()
    bg = cfg.get("bg_path", BG_PATH)
    if not os.path.exists(bg):
        if show_msg:
            QMessageBox.critical(None, "错误", f"背景图不存在:\n{bg}")
        return None

    img = Image.open(bg).convert("RGB")
    draw = ImageDraw.Draw(img)

    tc = tuple(cfg.get("text_color", [180, 0, 0]))
    dc = tuple(cfg.get("date_color", [100, 100, 100]))

    items = [
        (name, cfg["name"]["size"], cfg["name"]["x"], cfg["name"]["y"], tc, cfg["name"]["align"]),
        (name2, cfg["name2"]["size"], cfg["name2"]["x"], cfg["name2"]["y"], tc, cfg["name2"]["align"]),
        (name3, cfg["name3"]["size"], cfg["name3"]["x"], cfg["name3"]["y"], tc, cfg["name3"]["align"]),
        (school, cfg["school"]["size"], cfg["school"]["x"], cfg["school"]["y"], tc, cfg["school"]["align"]),
        (date, cfg["date"]["size"], cfg["date"]["x"], cfg["date"]["y"], dc, cfg["date"]["align"]),
    ]

    for text, size, x, y, color, align in items:
        font = get_font(size)
        bbox = draw.textbbox((0, 0), text, font=font)
        tw = bbox[2] - bbox[0]
        th = bbox[3] - bbox[1]
        px = x - tw // 2 if align == "center" else x
        py = y - th // 2 if align == "center" else y
        draw.text((px, py), text, font=font, fill=color)

    out_size = cfg.get("output_size", "original")
    if out_size != "original":
        w, h = map(int, out_size.split("x"))
        img = ImageOps.fit(img, (w, h), method=Image.LANCZOS)

    os.makedirs(cfg.get("output_dir", OUTPUT_DIR), exist_ok=True)
    fmt = cfg.get("output_format", "png")
    ext = "jpg" if fmt == "jpg" else "png"
    save_path = os.path.join(cfg["output_dir"], f"{name}_高考喜报.{ext}")
    save_kwargs = {"quality": cfg.get("quality", 95)} if ext == "jpg" else {}
    img.save(save_path, **save_kwargs)
    return save_path


def batch_from_excel(excel_path, progress_callback=None):
    wb = load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    results = []
    batch_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    idx = 0

    for i, row in enumerate(rows):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        name2 = str(row[1]).strip() if len(row) > 1 and row[1] else name
        name3 = str(row[2]).strip() if len(row) > 2 and row[2] else name2
        school = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        date = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        if not date:
            date = datetime.now().strftime("%Y年%m月%d日")

        path = create_card(name, name2, name3, school, date, show_msg=False)
        results.append((name, path))

        if progress_callback:
            progress_callback(i + 1, len(rows))

    return results


class SettingsDialog(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("位置与颜色设置")
        self.setMinimumWidth(500)
        self.cfg = load_config()
        self._build_ui()
        self._load_values()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        inner = QWidget()
        grid = QGridLayout(inner)
        grid.setVerticalSpacing(8)
        grid.setHorizontalSpacing(10)

        self.spin_name_x = QSpinBox()
        self.spin_name_x.setRange(0, 5000)
        self.spin_name_y = QSpinBox()
        self.spin_name_y.setRange(0, 5000)
        self.spin_name_size = QSpinBox()
        self.spin_name_size.setRange(10, 300)

        self.spin_school_x = QSpinBox()
        self.spin_school_x.setRange(0, 5000)
        self.spin_school_y = QSpinBox()
        self.spin_school_y.setRange(0, 5000)
        self.spin_school_size = QSpinBox()
        self.spin_school_size.setRange(10, 300)

        self.spin_date_x = QSpinBox()
        self.spin_date_x.setRange(0, 5000)
        self.spin_date_y = QSpinBox()
        self.spin_date_y.setRange(0, 5000)
        self.spin_date_size = QSpinBox()
        self.spin_date_size.setRange(10, 200)

        self.spin_name2_x = QSpinBox(); self.spin_name2_x.setRange(0, 5000)
        self.spin_name2_y = QSpinBox(); self.spin_name2_y.setRange(0, 5000)
        self.spin_name2_size = QSpinBox(); self.spin_name2_size.setRange(10, 200)

        self.spin_name3_x = QSpinBox(); self.spin_name3_x.setRange(0, 5000)
        self.spin_name3_y = QSpinBox(); self.spin_name3_y.setRange(0, 5000)
        self.spin_name3_size = QSpinBox(); self.spin_name3_size.setRange(10, 200)

        r = 0
        grid.addWidget(QLabel("<b>姓名</b>"), r, 0, 1, 6)
        r += 1
        grid.addWidget(QLabel("X:"), r, 0); grid.addWidget(self.spin_name_x, r, 1)
        grid.addWidget(QLabel("Y:"), r, 2); grid.addWidget(self.spin_name_y, r, 3)
        grid.addWidget(QLabel("字号:"), r, 4); grid.addWidget(self.spin_name_size, r, 5)

        r += 1
        grid.addWidget(QLabel("<b>姓名 (小字)</b>"), r, 0, 1, 6)
        r += 1
        grid.addWidget(QLabel("X:"), r, 0); grid.addWidget(self.spin_name2_x, r, 1)
        grid.addWidget(QLabel("Y:"), r, 2); grid.addWidget(self.spin_name2_y, r, 3)
        grid.addWidget(QLabel("字号:"), r, 4); grid.addWidget(self.spin_name2_size, r, 5)

        r += 1
        grid.addWidget(QLabel("<b>姓名 (小字2)</b>"), r, 0, 1, 6)
        r += 1
        grid.addWidget(QLabel("X:"), r, 0); grid.addWidget(self.spin_name3_x, r, 1)
        grid.addWidget(QLabel("Y:"), r, 2); grid.addWidget(self.spin_name3_y, r, 3)
        grid.addWidget(QLabel("字号:"), r, 4); grid.addWidget(self.spin_name3_size, r, 5)

        r += 1
        grid.addWidget(QLabel("<b>学校</b>"), r, 0, 1, 6)
        r += 1
        grid.addWidget(QLabel("X:"), r, 0); grid.addWidget(self.spin_school_x, r, 1)
        grid.addWidget(QLabel("Y:"), r, 2); grid.addWidget(self.spin_school_y, r, 3)
        grid.addWidget(QLabel("字号:"), r, 4); grid.addWidget(self.spin_school_size, r, 5)

        r += 1
        grid.addWidget(QLabel("<b>日期</b>"), r, 0, 1, 6)
        r += 1
        grid.addWidget(QLabel("X:"), r, 0); grid.addWidget(self.spin_date_x, r, 1)
        grid.addWidget(QLabel("Y:"), r, 2); grid.addWidget(self.spin_date_y, r, 3)
        grid.addWidget(QLabel("字号:"), r, 4); grid.addWidget(self.spin_date_size, r, 5)

        r += 1
        grid.addWidget(QLabel("<b>文字颜色</b> (RGB):"), r, 0, 1, 6)
        r += 1
        self.spin_r = QSpinBox(); self.spin_r.setRange(0, 255)
        self.spin_g = QSpinBox(); self.spin_g.setRange(0, 255)
        self.spin_b = QSpinBox(); self.spin_b.setRange(0, 255)
        grid.addWidget(QLabel("R:"), r, 0); grid.addWidget(self.spin_r, r, 1)
        grid.addWidget(QLabel("G:"), r, 2); grid.addWidget(self.spin_g, r, 3)
        grid.addWidget(QLabel("B:"), r, 4); grid.addWidget(self.spin_b, r, 5)

        r += 1
        grid.addWidget(QLabel("<b>输出设置</b>"), r, 0, 1, 6)
        r += 1
        self.combo_size = QComboBox()
        self.combo_size.addItems(PRESET_SIZES.keys())
        self.combo_format = QComboBox()
        self.combo_format.addItems(["png", "jpg"])
        grid.addWidget(QLabel("尺寸:"), r, 0); grid.addWidget(self.combo_size, r, 1, 1, 2)
        grid.addWidget(QLabel("格式:"), r, 3); grid.addWidget(self.combo_format, r, 4, 1, 2)

        scroll.setWidget(inner)
        layout.addWidget(scroll)

        btn_row = QHBoxLayout()
        btn_save = QPushButton("保存设置")
        btn_save.clicked.connect(self._save)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.close)
        btn_row.addStretch()
        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_cancel)
        layout.addLayout(btn_row)

    def _load_values(self):
        c = self.cfg
        self.spin_name_x.setValue(c["name"]["x"])
        self.spin_name_y.setValue(c["name"]["y"])
        self.spin_name_size.setValue(c["name"]["size"])
        self.spin_school_x.setValue(c["school"]["x"])
        self.spin_school_y.setValue(c["school"]["y"])
        self.spin_school_size.setValue(c["school"]["size"])
        self.spin_date_x.setValue(c["date"]["x"])
        self.spin_date_y.setValue(c["date"]["y"])
        self.spin_date_size.setValue(c["date"]["size"])
        self.spin_name2_x.setValue(c.get("name2", {"x": 527, "y": 740, "size": 48}).get("x", 527))
        self.spin_name2_y.setValue(c.get("name2", {"x": 527, "y": 740, "size": 48}).get("y", 740))
        self.spin_name2_size.setValue(c.get("name2", {"x": 527, "y": 740, "size": 48}).get("size", 48))
        self.spin_name3_x.setValue(c.get("name3", {"x": 527, "y": 650, "size": 40}).get("x", 527))
        self.spin_name3_y.setValue(c.get("name3", {"x": 527, "y": 650, "size": 40}).get("y", 650))
        self.spin_name3_size.setValue(c.get("name3", {"x": 527, "y": 650, "size": 40}).get("size", 40))
        tc = c.get("text_color", [180, 0, 0])
        self.spin_r.setValue(tc[0])
        self.spin_g.setValue(tc[1])
        self.spin_b.setValue(tc[2])

        out_size = c.get("output_size", "original")
        idx = list(PRESET_SIZES.values()).index(out_size) if out_size in PRESET_SIZES.values() else 0
        self.combo_size.setCurrentIndex(idx)
        fmt = c.get("output_format", "png")
        self.combo_format.setCurrentIndex(0 if fmt == "png" else 1)

    def _save(self):
        c = self.cfg
        c["name"]["x"] = self.spin_name_x.value()
        c["name"]["y"] = self.spin_name_y.value()
        c["name"]["size"] = self.spin_name_size.value()
        c["school"]["x"] = self.spin_school_x.value()
        c["school"]["y"] = self.spin_school_y.value()
        c["school"]["size"] = self.spin_school_size.value()
        c["date"]["x"] = self.spin_date_x.value()
        c["date"]["y"] = self.spin_date_y.value()
        c["date"]["size"] = self.spin_date_size.value()
        c.setdefault("name2", {"x": 527, "y": 740, "size": 48, "align": "center"})
        c.setdefault("name3", {"x": 527, "y": 650, "size": 40, "align": "center"})
        c["name2"]["x"] = self.spin_name2_x.value()
        c["name2"]["y"] = self.spin_name2_y.value()
        c["name2"]["size"] = self.spin_name2_size.value()
        c["name3"]["x"] = self.spin_name3_x.value()
        c["name3"]["y"] = self.spin_name3_y.value()
        c["name3"]["size"] = self.spin_name3_size.value()
        c["text_color"] = [self.spin_r.value(), self.spin_g.value(), self.spin_b.value()]
        key = self.combo_size.currentText()
        c["output_size"] = PRESET_SIZES.get(key, "original")
        c["output_format"] = self.combo_format.currentText()
        save_config(c)
        self.close()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("高考喜报自动生成系统")
        self.setMinimumSize(520, 620)
        self.cfg = load_config()
        self._build_ui()
        self._refresh_preview()

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)

        title = QLabel("<h2>🎓 高考喜报自动生成系统</h2>")
        title.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title)

        input_group = QGroupBox("输入信息")
        form = QGridLayout(input_group)
        form.setVerticalSpacing(8)

        self.input_name = QLineEdit()
        self.input_name.setPlaceholderText("请输入学生姓名")
        self.input_name.textChanged.connect(lambda: self._refresh_preview())

        self.input_school = QLineEdit()
        self.input_school.setPlaceholderText("请输入录取学校")
        self.input_school.textChanged.connect(lambda: self._refresh_preview())

        self.input_date = QLineEdit()
        self.input_date.setText(datetime.now().strftime("%Y年%m月%d日"))
        self.input_date.textChanged.connect(lambda: self._refresh_preview())

        form.addWidget(QLabel("姓  名:"), 0, 0)
        form.addWidget(self.input_name, 0, 1)
        form.addWidget(QLabel("学  校:"), 1, 0)
        form.addWidget(self.input_school, 1, 1)
        form.addWidget(QLabel("日  期:"), 2, 0)
        form.addWidget(self.input_date, 2, 1)

        main_layout.addWidget(input_group)

        btn_row1 = QHBoxLayout()
        self.btn_bg = QPushButton("选择背景图")
        self.btn_bg.clicked.connect(self._choose_bg)
        self.btn_font = QPushButton("选择字体")
        self.btn_font.clicked.connect(self._choose_font)
        self.btn_settings = QPushButton("位置/颜色设置")
        self.btn_settings.clicked.connect(self._open_settings)
        btn_row1.addWidget(self.btn_bg)
        btn_row1.addWidget(self.btn_font)
        btn_row1.addWidget(self.btn_settings)
        main_layout.addLayout(btn_row1)

        preview_group = QGroupBox("预览 (红点标记文字位置)")
        pv_layout = QVBoxLayout(preview_group)
        self.preview_label = QLabel()
        self.preview_label.setAlignment(Qt.AlignCenter)
        self.preview_label.setMinimumHeight(200)
        self.preview_label.setStyleSheet("background: #f0f0f0; border: 1px solid #ccc;")
        pv_layout.addWidget(self.preview_label)
        self.preview_info = QLabel("")
        self.preview_info.setAlignment(Qt.AlignCenter)
        pv_layout.addWidget(self.preview_info)
        main_layout.addWidget(preview_group)

        btn_row2 = QHBoxLayout()
        self.btn_generate = QPushButton("🎉 生成喜报")
        self.btn_generate.setMinimumHeight(36)
        self.btn_generate.setStyleSheet("font-weight:bold; font-size:14px;")
        self.btn_generate.clicked.connect(self._generate_single)

        self.btn_batch = QPushButton("📋 批量生成 (Excel)")
        self.btn_batch.setMinimumHeight(36)
        self.btn_batch.clicked.connect(self._generate_batch)

        btn_row2.addWidget(self.btn_generate)
        btn_row2.addWidget(self.btn_batch)
        main_layout.addLayout(btn_row2)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪 — 请填入信息后点击「生成喜报」")

    def _refresh_preview(self):
        bg = self.cfg.get("bg_path", BG_PATH)
        if not os.path.exists(bg):
            self.preview_label.setText("⚠ 未找到背景图\n请点击「选择背景图」")
            self.preview_info.setText("")
            return

        img = Image.open(bg).convert("RGB")
        w, h = img.size
        max_display_h = 250
        scale = max_display_h / h
        nw, nh = int(w * scale), int(h * scale)
        img_resized = img.resize((nw, nh), Image.LANCZOS)
        img_resized = img_resized.convert("RGBA")

        overlay = Image.new("RGBA", (nw, nh), (0, 0, 0, 0))
        odraw = ImageDraw.Draw(overlay)

        name_text = self.input_name.text() or "姓名位置"
        school_text = self.input_school.text() or "学校位置"
        date_text = self.input_date.text() or "日期位置"

        fc = tuple(self.cfg.get("text_color", [180, 0, 0]) + [180])
        dc = tuple(self.cfg.get("date_color", [100, 100, 100]) + [180])

        items = [
            (name_text, self.cfg["name"]["size"], self.cfg["name"]["x"], self.cfg["name"]["y"], fc),
            (school_text, self.cfg["school"]["size"], self.cfg["school"]["x"], self.cfg["school"]["y"], fc),
            (date_text, self.cfg["date"]["size"], self.cfg["date"]["x"], self.cfg["date"]["y"], dc),
        ]

        for text, size, x, y, color in items:
            sx = int(x * scale)
            sy = int(y * scale)
            ss = max(8, int(size * scale))
            r = max(4, ss // 6)
            odraw.ellipse([sx - r, sy - r, sx + r, sy + r], fill=(255, 0, 0, 200))
            odraw.text((sx + r + 2, sy - ss), text, fill=color)

        result = Image.alpha_composite(img_resized, overlay)
        data = result.tobytes("raw", "RGBA")
        qimage = QImage(data, nw, nh, QImage.Format_RGBA8888)
        pixmap = QPixmap.fromImage(qimage)
        self.preview_label.setPixmap(pixmap)
        self.preview_label.setFixedSize(nw + 4, nh + 4)

        filename = os.path.basename(bg)
        self.preview_info.setText(f"背景: {filename} | 尺寸: {w}x{h}")

    def _choose_bg(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择背景图片", "",
            "图片文件 (*.png *.jpg *.jpeg *.bmp);;所有文件 (*.*)"
        )
        if path:
            self.cfg["bg_path"] = path
            save_config(self.cfg)
            self.status_bar.showMessage(f"背景图已更新: {path}")
            self._refresh_preview()

    def _choose_font(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择字体文件", "",
            "字体文件 (*.ttf *.ttc *.otf);;所有文件 (*.*)"
        )
        if path:
            self.cfg["font_path"] = path
            save_config(self.cfg)
            self.status_bar.showMessage(f"字体已更新: {path}")

    def _open_settings(self):
        self.settings_dialog = SettingsDialog(self)
        self.settings_dialog.setWindowModality(Qt.ApplicationModal)
        self.settings_dialog.show()
        self.settings_dialog.destroyed.connect(lambda: self._on_settings_closed())

    def _on_settings_closed(self):
        self.cfg = load_config()
        self._refresh_preview()

    def _generate_single(self):
        name = self.input_name.text().strip()
        school = self.input_school.text().strip()
        date = self.input_date.text().strip()

        if not name:
            QMessageBox.warning(self, "提示", "请输入学生姓名")
            return
        if not school:
            QMessageBox.warning(self, "提示", "请输入录取学校")
            return

        self.cfg = load_config()
        path = create_card(name, name, name, school, date)
        if path:
            self.status_bar.showMessage(f"✅ 生成完成: {path}")
            QMessageBox.information(self, "生成成功", f"喜报已生成:\n{path}")

    def _generate_batch(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "",
            "Excel文件 (*.xlsx *.xls);;所有文件 (*.*)"
        )
        if not path:
            return

        self.cfg = load_config()

        progress = QProgressDialog("正在批量生成...", "取消", 0, 0, self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.show()

        all_results = []

        def callback(current, total):
            progress.setMaximum(total)
            progress.setValue(current)
            QApplication.processEvents()
            if progress.wasCanceled():
                return

        try:
            results = batch_from_excel(path, callback)
            all_results = results
        except Exception as e:
            QMessageBox.critical(self, "错误", f"Excel读取失败:\n{e}")
            progress.close()
            return

        progress.close()

        ok = sum(1 for _, p in all_results if p)
        fail = len(all_results) - ok
        msg = f"批量生成完成!\n成功: {ok} 张\n失败: {fail} 张"
        if ok:
            msg += f"\n\n保存位置: {self.cfg.get('output_dir', OUTPUT_DIR)}"
        QMessageBox.information(self, "批量生成结果", msg)
        self.status_bar.showMessage(f"批量生成完成: {ok} 张")


if __name__ == "__main__":
    app = QApplication([])

    font = QFont("Microsoft YaHei", 9)
    app.setFont(font)

    win = MainWindow()
    win.show()
    app.exec_()
